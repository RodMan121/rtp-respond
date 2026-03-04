"""
excel_export.py — Export de la matrice vers Excel
==================================================
Produit un fichier .xlsx avec :
- Onglet 1 : Matrice complète (toutes exigences)
- Onglet 2 : Must uniquement
- Onglet 3 : Exigences prédictives / implicites
- Onglet 4 : Synthèse par type
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from loguru import logger

# Palette couleurs MoSCoW
_MOSCOW_COLORS = {
    "must":   "C6EFCE",   # vert clair
    "should": "FFEB9C",   # jaune
    "could":  "DDEEFF",   # bleu clair
    "wont":   "F2DCDB",   # rouge clair
}

# Palette couleurs source
_SOURCE_COLORS = {
    "explicit":   "FFFFFF",
    "signal":     "FFF2CC",
    "implicit":   "EAF4FB",
    "predictive": "F3E5F5",
}

_HEADERS = [
    ("ID",               12),
    ("Type",             16),
    ("MoSCoW",           10),
    ("Source",           12),
    ("Exigence",         60),
    ("Citation source",  50),
    ("Partie prenante",  20),
    ("Justification",    45),
    ("Réponse proposée", 60),
]

def _header_row(ws, row=1):
    """Écrit la ligne d'en-tête avec mise en forme."""
    fill = PatternFill("solid", fgColor="2F4F7F")
    font = Font(bold=True, color="FFFFFF", size=10)
    for col, (label, _) in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28

def _set_col_widths(ws):
    for col, (_, width) in enumerate(_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

def _write_requirement(ws, row: int, item: dict):
    """Écrit une exigence sur une ligne."""
    moscow = item.get("moscow", "should")
    source = item.get("source", "explicit")
    bg = _SOURCE_COLORS.get(source, "FFFFFF")

    values = [
        item.get("id", ""),
        item.get("type", ""),
        moscow,
        source,
        item.get("text", ""),
        item.get("source_quote", ""),
        item.get("stakeholder", ""),
        item.get("rationale", ""),
        item.get("proposed_response", ""),
    ]

    moscow_fill = PatternFill("solid", fgColor=_MOSCOW_COLORS.get(moscow, "FFFFFF"))
    default_fill = PatternFill("solid", fgColor=bg)

    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(size=9)
        # Colonne MoSCoW colorée selon priorité
        if col == 3:
            cell.fill = moscow_fill
            cell.font = Font(size=9, bold=True)
        else:
            cell.fill = default_fill

    ws.row_dimensions[row].height = 45

def _write_sheet(ws, items: list[dict], title: str):
    ws.title = title
    _header_row(ws)
    for i, item in enumerate(items, start=2):
        _write_requirement(ws, i, item)
    _set_col_widths(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}{len(items)+1}"

def _summary_sheet(ws, matrix: list[dict]):
    ws.title = "Synthèse"
    headers = ["Type", "Must", "Should", "Could", "Won't", "Total"]
    fill_h = PatternFill("solid", fgColor="2F4F7F")
    font_h = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill_h
        c.font = font_h
        c.alignment = Alignment(horizontal="center")

    # Compter par type × MoSCoW
    types = sorted(set(x.get("type", "functional") for x in matrix))
    for row, t in enumerate(types, start=2):
        subset = [x for x in matrix if x.get("type") == t]
        ws.cell(row=row, column=1, value=t).font = Font(bold=True)
        for col, m in enumerate(["must", "should", "could", "wont"], start=2):
            ws.cell(row=row, column=col, value=sum(1 for x in subset if x.get("moscow") == m))
        ws.cell(row=row, column=6, value=len(subset))

    # Ligne totaux
    last = len(types) + 2
    ws.cell(row=last, column=1, value="TOTAL").font = Font(bold=True)
    for col, m in enumerate(["must", "should", "could", "wont"], start=2):
        ws.cell(row=last, column=col, value=sum(1 for x in matrix if x.get("moscow") == m)).font = Font(bold=True)
    ws.cell(row=last, column=6, value=len(matrix)).font = Font(bold=True)

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

def export(matrix: list[dict], output_path: str):
    wb = Workbook()

    # Onglet 1 : toutes exigences
    ws_all = wb.active
    _write_sheet(ws_all, matrix, "Matrice complète")

    # Onglet 2 : Must uniquement
    must_items = [x for x in matrix if x.get("moscow") == "must"]
    ws_must = wb.create_sheet()
    _write_sheet(ws_must, must_items, "Must")

    # Onglet 3 : exigences implicites + prédictives
    augmented = [x for x in matrix if x.get("source") in ("implicit", "predictive", "signal", "stakeholders", "signals")]
    ws_aug = wb.create_sheet()
    _write_sheet(ws_aug, augmented, "Implicites & Prédictives")

    # Onglet 4 : synthèse
    ws_sum = wb.create_sheet()
    _summary_sheet(ws_sum, matrix)

    wb.save(output_path)
    logger.success(f"📊 Export Excel : {output_path} ({len(matrix)} exigences)")
